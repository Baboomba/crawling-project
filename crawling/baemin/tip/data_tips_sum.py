
import datetime
import os
from openpyxl import load_workbook
from openpyxl import workbook, worksheet
import pandas as pd
import sys
sys.path.append(r'.\crawling')

from common.Database import DataProcess



class TipCalculation(DataProcess):
    def __init__(self, kind='tip'):
        super().__init__(kind)
        self.file_dir = r'C:\Users\SEC\Coding\VScode\crawling\result\202211_tips'
        self.save_dir = r'C:\Users\SEC\Coding\VScode\crawling\result\202211_tips\_tip.xlsx'
        
        
    def files_into_list(self):
        file_list = os.listdir(self.file_dir) # getting the name of files
        return file_list
    
    
    def file_range(self):
        size = len(self.files_into_list())
        return size
    
    
    def tip_extraction(self):
        file_list = self.files_into_list()
        store_list = []
        
        for index, name in enumerate(file_list):
            file_name = rf'C:\Users\SEC\Coding\VScode\crawling\result\202211_tips\{name}'
            wb = load_workbook(file_name, data_only=True)
            ws = wb['상세']

            tip = []

            for num, item in enumerate(ws[5]):
                    if item.value == '바로결제배달팁':
                        col_a = num + 1
                    elif item.value == '만나서결제배달팁':
                        col_b = num + 1

            for m in range(6, 200):    # adding the details
                col1 = ws.cell(row=m, column=col_a).value
                col2 = ws.cell(row=m, column=col_b).value
                if col1 or col2 != None:
                    col3 = col1 + col2
                tip.append(col3)

            tip_sum = list(filter(None, tip))
            tip_sum = sum(tip_sum)
            store = [index, name, tip_sum]
            store_list.append(store)
            
        return store_list
            
            
    def tip_framed(self):
        store_list = self.tip_extraction()
        size = self.file_range()
        frame_tip = pd.DataFrame(store_list,
                                 index=range(size),
                                 columns=['No.', 'store', 'tip'])
        return frame_tip
    
    
    def save_tips(self):
        frame_tip = self.tip_framed()
        frame_tip.to_excel(self.save_dir)
    

if __name__ == '__main__':
    print(datetime.datetime.now())
    tip = TipCalculation()
    tip.save_tips()
    print('complete')
    print(datetime.datetime.now())