import datetime
import pandas as pd

from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import sys
sys.path.append(r'./')


class SaveData():
    def data_dir(self, app, kind, period):
        yesterday = datetime.datetime.today() - datetime.timedelta(days=1)
        day = yesterday.strftime('%Y%m%d')
        month = int(datetime.datetime.today().strftime('%m')) - 1
        if period == 'day':
            dir = rf'./result/{app}_{kind}_{day}.xlsx'
        elif period == 'month':
            dir = rf'./result/{app}_{kind}_{month}.xlsx'
        return dir
    
    
    def save_data(self, result, app, kind, period):
        data_dir = self.data_dir(app, kind, period)
        result.to_excel(data_dir)
        
    
    def set_index(self, result):
        result.set_index('No.', drop=True, inplace=True)
        result.fillna("", inplace=True)
        return result
    
    
    
    def process_data(self):
        wb = load_workbook(self.data_dir)
        ws = wb.active
        ws.column_dimensions[get_column_letter(1)].width = 3.5
        ws.column_dimensions[get_column_letter(2)].width = 11.5
        ws.column_dimensions[get_column_letter(3)].width = 11.5
        ws.column_dimensions[get_column_letter(4)].width = 4
        ws.column_dimensions[get_column_letter(5)].width = 60
        ws.column_dimensions[get_column_letter(6)].width = 11.5
        ws.column_dimensions[get_column_letter(7)].width = 4.5
        ws.column_dimensions[get_column_letter(8)].width = 30
        
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 150
        
        for order in range(0, ws.max_column):
            ws[row][order].alignment = Alignment(horizontal = 'center', vertical='center', wrap_text=True)
    
        for image_rows in range(2, ws.max_row + 1):
            img_no = ws[image_rows][6].value
            try:
                review_img = Image(r'./download/baemin_img/{}.jpg'.format(img_no))
                review_img.height = 200
                review_img.width = 200
                ws.add_image(review_img, 'H{}'.format(image_rows))
            except:
                continue
        
        wb.save(self.data_dir)