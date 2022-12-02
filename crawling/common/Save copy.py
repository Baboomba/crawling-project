import datetime
import pandas as pd

from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import sys
sys.path.append(r'.\crawling')

from common.Database import DataProcess


class SaveData(DataProcess):
    def __init__(self, app, kind):
        yesterday = datetime.datetime.today() - datetime.timedelta(days=1)
        day = yesterday.strftime('%Y%m%d')
        month = int(datetime.datetime.today().strftime('%m')) - 1
        self.data_dir_day = r'.\crawling\result\{}_{}_{}.xlsx'.format(app, kind, day)
        self.data_dir_month = r'.\crawling\result\{}_{}_{}.xlsx'.format(app, kind, month)
        self.data_dir_tip = r'.\crawling\result\{}_{}_{}.xlsx'.format(app, kind, month)
    
    
    def sales_day(self, result):
        result.to_excel(self.data_dir_day)        
    
    
    def sales_month(self, store_index, df_result):
        global df_month
        
        if store_index == 0:
            df_month = self.frame_sales
        
        df_month = pd.concat([df_month, df_result])
        df_month.to_excel(self.data_dir_month)
        return df_month
    
    
    def tip_bm(self, store_index, df_result):
        global df_tips
        
        if store_index == 0:
            df_tips = self.frame_tips
        
        df_tips = pd.concat([df_tips, df_result])
        df_tips.to_excel(self.data_dir_tip)
        return df_tips
    
    
    def review_BM(self, store_index, df_result):
        global df_review
        
        if store_index == 0:
            df_review = self.frame_review
                    
        df_review = pd.concat([df_review, df_result])
        df_review.fillna('', inplace=True)
        df_review.to_excel(self.data_dir)
        return df_review
    
    
    def process_data(self):
        wb = load_workbook(self.data_dir)
        ws = wb.active
        ws.column_dimensions[get_column_letter(1)].width = 3.5
        ws.column_dimensions[get_column_letter(2)].width = 11.5
        ws.column_dimensions[get_column_letter(3)].width = 11.5
        ws.column_dimensions[get_column_letter(4)].width = 4
        ws.column_dimensions[get_column_letter(5)].width = 60
        ws.column_dimensions[get_column_letter(6)].width = 11.5
        ws.column_dimensions[get_column_letter(7)].width = 4.5
        ws.column_dimensions[get_column_letter(8)].width = 30
        
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 150
        
        for order in range(0, ws.max_column):
            ws[row][order].alignment = Alignment(horizontal = 'center', vertical='center', wrap_text=True)
    
        for image_rows in range(2, ws.max_row + 1):
            img_no = ws[image_rows][6].value
            try:
                review_img = Image(r'.\crawling\download\baemin_img\{}.jpg'.format(img_no))
                review_img.height = 200
                review_img.width = 200
                ws.add_image(review_img, 'H{}'.format(image_rows))
            except:
                continue
        
        wb.save(self.data_dir)


    def set_index(self):
        df_sales.set_index('No.', drop=True, inplace=True)
        df_sales.fillna("", inplace=True)
        df_sales.to_excel(self.data_dir)

