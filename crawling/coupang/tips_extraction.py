# extract store names in coupang files

#import pandas as pd
#import openpyxl as px

#store_name = []
#df = pd.DataFrame(store_name)


#for x in range(0, 66):
#    files = r'C:\Users\SEC\Downloads\cou1\coupang_eats_2022-10 ({}).xlsx'.format(x)
#    wb = px.load_workbook(files, data_only=True)
#    ws = wb.active
#    names = ws["G4"].value
#    store_name.append(names)

#df['name'] = store_name    

#df.to_excel(r'C:\Users\SEC\Downloads\store.xlsx')

#print(df)


# extracting tips of coupang's files

import pandas as pd
import openpyxl as px


df = pd.DataFrame(index=range(120), columns=['store', 'tips'])

for x in range(0, 118):
    files = rf'C:\Users\SEC\Coding\VScode\crawling\download\coupnag_tip\coupang_eats_2022-11 ({x}).xlsx'
    wb = px.load_workbook(files)
    ws = wb.active
    tips = []
    
    for i in range(2, len(ws['U'])):
        a_tip = ws['U'][i].value
        tips.append(a_tip)
    tips = sum(tips)
    df['store'][x] = ws["G4"].value
    df['tips'][x] = tips
    print(df)

df.to_excel(r'C:\Users\SEC\Coding\VScode\crawling\download\coupang_tip.xlsx')


